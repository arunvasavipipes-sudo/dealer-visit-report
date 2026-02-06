import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

st.set_page_config(page_title="Dealer Visit Report Generator")

st.title("Upload Excel → Download Final Report")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

def extract_amount(text):
    if pd.isna(text):
        return None
    nums = re.findall(r"\d[\d,]*", str(text))
    return int(nums[0].replace(",", "")) if nums else None

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    df = df.dropna(axis=1, how="all")

    df = df.rename(columns={
        "Party Name": "Party",
        "Place": "Place",
        "Date": "Date",
        "Time": "Time",
        "Order value": "Order",
        "Visited Purpose": "Purpose",
        "Name of the Executive": "Executive"
    })

    if st.button("Generate Final Report"):
        output = []
        row_pointer = 0

        for exec_name, g in df.groupby("Executive"):
            g = g.sort_values("Time")

            clock_in = g["Time"].iloc[0]
            clock_out = g["Time"].iloc[-1]

            output.append(["VASAVI PIPES PVT LTD GUNTAKAL"])
            output.append([f"{exec_name} (Clock in {clock_in}, Clock out {clock_out})"])
            output.append(["S.No", "Date", "Party Name", "Place", "Time", "Order Value", "Visited Purpose"])

            total = 0
            for i, r in enumerate(g.itertuples(), start=1):
                amt = extract_amount(r.Order)
                total += amt if amt else 0

                output.append([
                    i,
                    r.Date.date(),
                    r.Party,
                    r.Place,
                    r.Time,
                    amt if amt else "–",
                    r.Purpose
                ])

            output.append(["", "", "", "", "TOTAL", total, ""])
            output.append([""])

        final_df = pd.DataFrame(output)
        final_df.to_excel("Final_Report.xlsx", index=False, header=False)

        wb = load_workbook("Final_Report.xlsx")
        ws = wb.active

        bold = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value in [
                    "VASAVI PIPES PVT LTD GUNTAKAL",
                    "S.No"
                ] or str(cell.value).startswith("TOTAL"):
                    cell.font = bold
                if cell.row > 2 and cell.column <= 7:
                    cell.border = border
                cell.alignment = Alignment(vertical="center")

        wb.save("Final_Report.xlsx")

        with open("Final_Report.xlsx", "rb") as f:
            st.download_button("Download Final Excel", f, file_name="Visited_Dealers_Report.xlsx")


